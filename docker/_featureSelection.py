#!/usr/bin/env python
import matplotlib as mpl
mpl.use('pdf')
from time import time
import numpy as np
import pandas as pd 
import argparse
import glob
import shutil
import os
from  openpyxl import Workbook
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.linear_model import RandomizedLasso
from sklearn.ensemble import RandomForestRegressor
from sklearn.cross_validation import cross_val_score, ShuffleSplit
from sklearn.linear_model import Lasso
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import Ridge
from sklearn.metrics import r2_score
from sklearn.feature_selection import RFE
from sklearn.linear_model import LinearRegression
np.random.seed(42)







def corplot(X,filesavename='all.pdf'):
	corr = X.corr()
	cmap = sns.diverging_palette(220, 10, as_cmap=True)
	# Generate a mask for the upper triangle
	mask = np.zeros_like(corr, dtype=np.bool)
	mask[np.triu_indices_from(mask)] = True
	fig,ax = plt.subplots(figsize=(22, 18))
	ax.set_title('Feature Correlation',fontweight='bold')
	sns.heatmap(corr, mask=mask, cmap=cmap, vmax=.7,square=True,linewidths=.5,annot=True, cbar_kws={"shrink": .5}, ax=ax	)
	# Generate a custom diverging colormap
	fig.tight_layout()
	plt.savefig(filesavename, dpi = 300)

def featureselection(datset,output='results'):
	t0 = time()
	dataset = pd.read_csv(datset)
	dataset.to_csv(output[:-4]+'.csv')
	# DO feature evaluation and write a xlsx file. 
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "feature selection scores"
	rownum=2
	ws1.cell(column=1, row=1).value='Feature name'
	ws1.cell(column=2, row=1).value='Stability Selection'
	ws1.cell(column=3, row=1).value='Univariate using random forest regressor (r2 measure)'
	ws1.cell(column=4, row=1).value='Univariate using random forest regressor (auc measure)'
	ws1.cell(column=5, row=1).value='L1 regularization / Lasso'
	ws1.cell(column=6, row=1).value='L3 regularization / Ridge'
	ws1.cell(column=7, row=1).value="Mean decrease impurity"
	ws1.cell(column=8, row=1).value='Recursive feature elimination'
	Collumnheadeers=list(dataset.columns.values)
	for imagebiom in Collumnheadeers:
		print imagebiom
		ws1.cell(column=1, row=rownum).value=imagebiom
		rownum+=1
	print dataset
	print Collumnheadeers
	# Create and save correlation plots. 
	## Get labels
	y=dataset['label'].values
	## Delete labes from list 
	Collumnheadeers.remove('label')
	X=dataset[Collumnheadeers]
	corplot(X,filesavename='all.pdf')	
	# performe feature selection
	rlasso = RandomizedLasso(alpha=0.00025)
	rlasso.fit(X, y)
	print "Features sorted by their score:"
	print sorted(zip(map(lambda x: round(x, 4), rlasso.scores_), 
					 Collumnheadeers), reverse=True)
	rownum=2
	for score_val in rlasso.scores_.tolist():
		ws1.cell(column=2, row=rownum).value=score_val
		rownum+=1
		rlasso.scores_
	print np.where(rlasso.scores_ > 0.8)[0] + 1
	print Collumnheadeers
	elementselect=np.where(rlasso.scores_ > 0.8)[0] 
	Collumnheadeersel=[]
	for i in elementselect:
		Collumnheadeersel.append(Collumnheadeers[i])
	corplot(X[Collumnheadeersel],filesavename='selectedstability.pdf')	
	wb.save(filename = 'ResultTableIndividualFeature.xlsx')


	rf = RandomForestRegressor(n_estimators=20, max_depth=4)
	scores = []
	scoresval= []
	X1=X.as_matrix()
	for i in range(X.shape[1]):
		 score = cross_val_score(rf, X1[:, i:i+1], y, scoring="r2",
								  cv=ShuffleSplit(len(X), 3, .3))
		 scores.append((round(np.mean(score), 3), Collumnheadeers[i]))
		 scoresval.append(round(np.mean(score), 3))
	print sorted(scores, reverse=True)
	rownum=2
	for score_val in scoresval:
		ws1.cell(column=3, row=rownum).value=score_val
		rownum+=1
	scoresval= []	
	for i in range(X.shape[1]):
		 score = cross_val_score(rf, X1[:, i:i+1], y, scoring="roc_auc",
								  cv=ShuffleSplit(len(X), 3, .3))
		 scores.append((round(np.mean(score), 3), Collumnheadeers[i]))
		 scoresval.append(round(np.mean(score), 3))
	print sorted(scores, reverse=True)
	rownum=2
	for score_val in scoresval:
		ws1.cell(column=4, row=rownum).value=score_val
		rownum+=1
	scaler = StandardScaler()
	X3 = scaler.fit_transform(X1)
	lasso = Lasso(alpha=.0003)
	lasso.fit(X3, y)
	print "Features sorted by their score:"
	print sorted(zip(map(lambda x: round(x, 4), lasso.coef_), 
					 Collumnheadeers), reverse=True)
	rownum=2
	for score_val in lasso.coef_.tolist():
		ws1.cell(column=5, row=rownum).value=score_val
		rownum+=1

	scaler = StandardScaler()
	X3 = scaler.fit_transform(X1)
	ridge = Ridge(alpha=10)
	ridge.fit(X3,y)
	print "Features sorted by their score:"
	print sorted(zip(map(lambda x: round(x, 4), lasso.coef_), 
					 Collumnheadeers), reverse=True)
	rownum=2
	for score_val in lasso.coef_.tolist():
		ws1.cell(column=6, row=rownum).value=score_val
		rownum+=1

	rf = RandomForestRegressor()
	rf.fit(X, y)
	rownum=2
	for score_val in rf.feature_importances_.tolist():
		ws1.cell(column=7, row=rownum).value=score_val
		rownum+=1
	#use linear regression as the model
	lr = LinearRegression()
	#rank all features, i.e continue the elimination until the last one
	rfe = RFE(lr, n_features_to_select=1)
	rfe.fit(X,y)
	rownum=2
	for score_val in rfe.ranking_.tolist():
		ws1.cell(column=8, row=rownum).value=score_val
		rownum+=1
	wb.save(filename = 'ResultTableIndividualFeature.xlsx')
	path_=os.getcwd()
	directory=path_+'/output/'
	if not os.path.exists(directory):
		os.makedirs(directory)

	types = ('*.pdf', '*.csv','*.xlsx') # the tuple of file types
	files_grabbed = []
	for files in types:
		files_grabbed.extend(glob.glob(files))
	for file in files_grabbed:
		if os.path.isfile(file):
			shutil.copy2(file, directory)	
	shutil.make_archive(output[:-4], 'zip', directory)
	return 0


def main(argv):
	featureselection(argv.datset, argv.output)
	return 0

if __name__ == "__main__":
	parser = argparse.ArgumentParser( description='Machine learning analysis pipeline')
	parser.add_argument ("-i", "--datset",  help="dataset (csv type)" , required=True)
	parser.add_argument ("-o", "--output",  help="output name of zip file" , required=True)
	parser.add_argument('--version', action='version', version='%(prog)s 0.1')
	parser.add_argument("-q", "--quiet",
						  action="store_false", dest="verbose",
						  default=True,
						  help="don't print status messages to stdout")
	args = parser.parse_args()
	main(args)