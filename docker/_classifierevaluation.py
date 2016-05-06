#!/usr/bin/env python
import matplotlib as mpl
mpl.use('pdf')
from time import time
import numpy as np
import pandas as pd 
from sklearn.cluster import KMeans
import argparse
import nibabel as nib
from sklearn.preprocessing import StandardScaler
import pandas as pd
import seaborn as sns
from sklearn.linear_model import RandomizedLasso
import matplotlib.pyplot as plt
from sklearn.cross_validation import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import LinearSVC
from sklearn.calibration import calibration_curve
from sklearn.cross_validation import StratifiedShuffleSplit
from sklearn.grid_search import GridSearchCV
from  openpyxl import Workbook
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler
from _analyticscalc import analyticscalc
from sklearn.learning_curve import learning_curve
from sklearn.cross_validation import KFold
from sklearn import cross_validation
from sklearn.learning_curve import learning_curve
from sklearn.metrics import classification_report
from sklearn.metrics import confusion_matrix
import glob
import shutil
import os

np.random.seed(42)


def plot_learning_curve(estimator, title, X, y, ylim=None, cv=None,
                        n_jobs=1, train_sizes=np.linspace(.1, 1.0, 5)):
    """
    Generate a simple plot of the test and traning learning curve.

    Parameters
    ----------
    estimator : object type that implements the "fit" and "predict" methods
        An object of that type which is cloned for each validation.

    title : string
        Title for the chart.

    X : array-like, shape (n_samples, n_features)
        Training vector, where n_samples is the number of samples and
        n_features is the number of features.

    y : array-like, shape (n_samples) or (n_samples, n_features), optional
        Target relative to X for classification or regression;
        None for unsupervised learning.

    ylim : tuple, shape (ymin, ymax), optional
        Defines minimum and maximum yvalues plotted.

    cv : integer, cross-validation generator, optional
        If an integer is passed, it is the number of folds (defaults to 3).
        Specific cross-validation objects can be passed, see
        sklearn.cross_validation module for the list of possible objects

    n_jobs : integer, optional
        Number of jobs to run in parallel (default 1).
    """
    plt.figure()
    plt.title(title)
    if ylim is not None:
        plt.ylim(*ylim)
    plt.xlabel("Training examples")
    plt.ylabel("Score")
    train_sizes, train_scores, test_scores = learning_curve(
        estimator, X, y, cv=cv, n_jobs=n_jobs, train_sizes=train_sizes)
    train_scores_mean = np.mean(train_scores, axis=1)
    train_scores_std = np.std(train_scores, axis=1)
    test_scores_mean = np.mean(test_scores, axis=1)
    test_scores_std = np.std(test_scores, axis=1)
    plt.grid()

    plt.fill_between(train_sizes, train_scores_mean - train_scores_std,
                     train_scores_mean + train_scores_std, alpha=0.1,
                     color="r")
    plt.fill_between(train_sizes, test_scores_mean - test_scores_std,
                     test_scores_mean + test_scores_std, alpha=0.1, color="g")
    plt.plot(train_sizes, train_scores_mean, 'o-', color="r",
             label="Training score")
    plt.plot(train_sizes, test_scores_mean, 'o-', color="g",
             label="Cross-validation score")

    plt.legend(loc="best")
    return plt


def plot_confusion_matrix(y_test, pred):
	labels = ['Class1', 'Class1']
	cm = confusion_matrix(y_test, pred, labels)
	print(cm)
	fig = plt.figure()
	ax = fig.add_subplot(111)
	cax = ax.matshow(cm)
	plt.title('Confusion matrix of the classifier')
	fig.colorbar(cax)
	ax.set_xticklabels([''] + labels)
	ax.set_yticklabels([''] + labels)
	plt.xlabel('Predicted')
	plt.ylabel('True')
	plt.show()




def calibrationplot(X_train,X_test,y_train,y_test,filesavename='calibrationplot.pdf'):
	# Create classifiers
	lr = LogisticRegression()
	gnb = GaussianNB()
	svc = LinearSVC(C=1.0)
	rfc = RandomForestClassifier(n_estimators=100)
	###############################################################################
	# Plot calibration plots

	plt.figure(figsize=(10, 10))
	ax1 = plt.subplot2grid((3, 1), (0, 0), rowspan=2)
	ax2 = plt.subplot2grid((3, 1), (2, 0))

	ax1.plot([0, 1], [0, 1], "k:", label="Perfectly calibrated")
	for clf, name in [(lr, 'Logistic'),
					  (gnb, 'Naive Bayes'),
					  (svc, 'Support Vector Classification'),
					  (rfc, 'Random Forest')]:
		clf.fit(X_train, y_train)
		if hasattr(clf, "predict_proba"):
			prob_pos = clf.predict_proba(X_test)[:, 1]
		else:  # use decision function
			prob_pos = clf.decision_function(X_test)
			prob_pos = \
				(prob_pos - prob_pos.min()) / (prob_pos.max() - prob_pos.min())
		fraction_of_positives, mean_predicted_value = \
			calibration_curve(y_test, prob_pos, n_bins=10)

		ax1.plot(mean_predicted_value, fraction_of_positives, "s-",
				 label="%s" % (name, ))

		ax2.hist(prob_pos, range=(0, 1), bins=10, label=name,
				 histtype="step", lw=2)

	ax1.set_ylabel("Fraction of positives")
	ax1.set_ylim([-0.05, 1.05])
	ax1.legend(loc="lower right")
	ax1.set_title('Calibration plots  (reliability curve)')

	ax2.set_xlabel("Mean predicted value")
	ax2.set_ylabel("Count")
	ax2.legend(loc="upper center", ncol=2)

	plt.tight_layout()
	plt.savefig(filesavename, dpi = 300)
def corplot(X,filesavename='all.pdf'):
	corr = X.corr()
	cmap = sns.diverging_palette(220, 10, as_cmap=True)
	# Generate a mask for the upper triangle
	mask = np.zeros_like(corr, dtype=np.bool)
	mask[np.triu_indices_from(mask)] = True
	fig,ax = plt.subplots(figsize=(22, 18))
	ax.set_title('Feature Correlation',fontweight='bold')
	sns.heatmap(corr, mask=mask, cmap=cmap, vmax=.7,square=True,linewidths=.5,annot=True, cbar_kws={"shrink": .5}, ax=ax	)
	# Generate a custom diverging colormap
	fig.tight_layout()
	plt.savefig(filesavename, dpi = 300)
def machinelearningpipeline(datset,output='results.zip'):
	t0 = time()
	dataset = pd.read_csv(datset)
	dataset.to_csv(output[:-4]+'.csv')
	# DO feature evaluation and write a xlsx file. 
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "ResultTableIndividualFeature"
	rownum=2
	ws1.cell(column=1, row=1).value='Feature name'
	ws1.cell(column=2, row=1).value='Az'
	ws1.cell(column=3, row=1).value='Optimal threshold'
	ws1.cell(column=4, row=1).value='Sensitivity'
	ws1.cell(column=5, row=1).value='Specificity'
	ws1.cell(column=6, row=1).value='Confidence interval: low'
	ws1.cell(column=7, row=1).value='Confidence interval: high'
	Collumnheadeers=list(dataset.columns.values)
	for imagebiom in Collumnheadeers:
		ValuesMetric= dataset[imagebiom].values
		Targets= dataset['label'].values
		roc_auc_score, optimalval, sens,spec, confidence_lower, confidence_upper=analyticscalc(ValuesMetric,Targets,imagebiom)
		ws1.cell(column=1, row=rownum).value=imagebiom
		ws1.cell(column=2, row=rownum).value="{:0.3f}".format(roc_auc_score)
		ws1.cell(column=3, row=rownum).value="{:0.3f}".format(optimalval)
		ws1.cell(column=4, row=rownum).value="{:0.3f}".format(sens)
		ws1.cell(column=5, row=rownum).value="{:0.3f}".format(spec)
		ws1.cell(column=6, row=rownum).value="{:0.3f}".format(confidence_lower)
		ws1.cell(column=7, row=rownum).value="{:0.3f}".format(confidence_upper)
		rownum+=1
	wb.save(filename = 'ResultTableIndividualFeature.xlsx')
	print dataset
	print Collumnheadeers
	# Create and save correlation plots. 
	## Get labels
	y=dataset['label'].values
	## Delete labes from list 
	Collumnheadeers.remove('label')
	X=dataset[Collumnheadeers]
	corplot(X,filesavename='all.pdf')	
	# performe feature selection
	rlasso = RandomizedLasso(alpha=0.00025)
	rlasso.fit(X, y)
	print "Features sorted by their score:"
	print sorted(zip(map(lambda x: round(x, 4), rlasso.scores_), 
					 Collumnheadeers), reverse=True)
	print np.where(rlasso.scores_ > 0.8)[0] + 1
	print Collumnheadeers
	elementselect=np.where(rlasso.scores_ > 0.8)[0] 
	Collumnheadeersel=[]
	for i in elementselect:
		Collumnheadeersel.append(Collumnheadeers[i])
	corplot(X[Collumnheadeersel],filesavename='selected.pdf')	
	# optimize and evalute classifier 
	X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0)
	# Compare classifier 
	calibrationplot(X_train,X_test,y_train,y_test,filesavename='calibrationplot.pdf')
	X=X_train.copy()
	y=y_train.copy()
	# Run a quick example on non optimal classifiers
	clf1 = LogisticRegression()
	clf2 = RandomForestClassifier()
	clf3 = GaussianNB()
	clf4= SVC()
	print('5-fold cross validation:\n')
	for clf, label in zip([clf1, clf2, clf3, clf4], ['Logistic Regression', 'Random Forest', 'naive Bayes','SVM']):
		scores = cross_validation.cross_val_score(clf, X, y, cv=5, scoring='roc_auc', n_jobs=1)
		print("roc_auc: %0.2f (+/- %0.2f) [%s]" % (scores.mean(), scores.std(), label))
	# SVM
	scaler = StandardScaler()
	X1 = scaler.fit_transform(X)
	C_range = np.logspace(-2, 10, 13)
	gamma_range = np.logspace(-9, 3, 13)
	param_grid = dict(gamma=gamma_range, C=C_range)
	cv = StratifiedShuffleSplit(y, n_iter=5, test_size=0.2, random_state=42)
	grid = GridSearchCV(SVC(kernel='rbf'), param_grid=param_grid, cv=cv,scoring='roc_auc')
	grid.fit(X1, y)
	scores = [x[1] for x in grid.grid_scores_]
	scores = np.array(scores).reshape(len(C_range), len(gamma_range))
	# Draw heatmap of the validation accuracy as a function of gamma and C
	plt.figure(figsize=(8, 6))
	plt.subplots_adjust(left=.2, right=0.95, bottom=0.15, top=0.95)
	plt.imshow(scores, interpolation='nearest', cmap=plt.cm.jet)
	plt.xlabel('gamma')
	plt.ylabel('C')
	plt.colorbar()
	plt.xticks(np.arange(len(gamma_range)), gamma_range, rotation=45)
	plt.yticks(np.arange(len(C_range)), C_range)
	plt.title('Validation accuracy')
	plt.savefig('SVMheatmap.pdf', dpi = 300)
	print("The best parameters are %s with a score of %0.2f"
		  % (grid.best_params_, grid.best_score_))

	# Check out random forest accuracy...
	scores = ['roc_auc']#['precision_weighted', 'recall_weighted','roc_auc']
	Random_plot=[]
	tuned_parameters = [{'n_estimators': [1,10,20,30,40,50,60,70,80,90,100,200,300,400,500,600,700,800,900,1000]}]
	for score in scores:
		print("# Tuning hyper-parameters for %s" % score)
		print()

		clf = GridSearchCV(RandomForestClassifier(), tuned_parameters, cv=5,n_jobs=40,
						   scoring='%s' % score)
		clf.fit(X, y)
		print("Best parameters set found on development set:")
		print()
		print(clf.best_params_)
		print()
		print("Grid scores on development set:")
		print()
		for params, mean_score, scores in clf.grid_scores_:
			print("%0.3f (+/-%0.03f) for %r"
				  % (mean_score, scores.std() * 2, params))
			Random_plot.append(mean_score)
	param=clf.best_params_

	f, ax = plt.subplots(figsize=(20, 20))
	plt.plot( [1,10,20,30,40,50,60,70,80,90,100,200,300,400,500,600,700,800,900,1000],Random_plot,  lw=2)
	plt.title("The best parameter is n_estimators=%s with area under ROC of %0.2f"
		  % (param.get("n_estimators"), clf.best_score_),fontweight='bold')
	plt.xlabel('Numer of estimators',fontweight='bold')
	plt.ylabel('Area Under ROC (Az)',fontweight='bold')
	plt.savefig('RandomForrest.pdf', tight_layout=True, dpi = 600)
	f, ax = plt.subplots(figsize=(20, 20))
	title = 'Learning Curves (SVM)' 
	param=grid.best_params_
	estimator = SVC(kernel='rbf',C=param.get("C"), gamma= param.get("gamma"))
	print y_train
	cv = cross_validation.ShuffleSplit(X.shape[0], n_iter=10,
                                   test_size=0.2, random_state=0)
	plot_learning_curve(estimator, title, X, y, (0.7, 1.01), cv=cv, n_jobs=4)
	plt.savefig('LearningCurvesSVM.pdf', tight_layout=True, dpi = 600)
	print str((time() - t0))
	estimator.fit(X, y)
	y_true, y_pred = y_test, estimator.predict(X_test)
	print(classification_report(y_true, y_pred))
	cm = confusion_matrix(y_test, y_pred)
	np.set_printoptions(precision=2)
	print('Confusion matrix, without normalization')
	print(cm)
	plt.figure()
	# Normalize the confusion matrix by row (i.e by the number of samples
	# in each class)
	cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
	print('Normalized confusion matrix')
	print(cm_normalized)
	plt.figure()
	plot_confusion_matrix(y_test, y_pred)
	plt.savefig('ConfusionMatrixSVM.pdf', tight_layout=True, dpi = 600)
	# os.chdir(os.getcwd())
	# for file in glob.glob("*.pdf"):
	#     print(file)
	# shutisl.make_archive(output_filename, 'zip', dir_name)
	path_=os.getcwd()
	directory=path_+'/output/'
	if not os.path.exists(directory):
	    os.makedirs(directory)

	types = ('*.pdf', '*.csv','*.xlsx') # the tuple of file types
	files_grabbed = []
	for files in types:
	    files_grabbed.extend(glob.glob(files))
	for file in files_grabbed:
	    if os.path.isfile(file):
	        shutil.copy2(file, directory)	
	shutil.make_archive(output[:-4], 'zip', directory)
	return 0


def main(argv):
	machinelearningpipeline(argv.datset, argv.output)
	return 0

if __name__ == "__main__":
	parser = argparse.ArgumentParser( description='Machine learning analysis pipeline')
	parser.add_argument ("-i", "--datset",  help="dataset (csv type)" , required=True)
	parser.add_argument ("-o", "--output",  help="output name of zip file" , required=True)
	parser.add_argument('--version', action='version', version='%(prog)s 0.1')
	parser.add_argument("-q", "--quiet",
						  action="store_false", dest="verbose",
						  default=True,
						  help="don't print status messages to stdout")
	args = parser.parse_args()
	main(args)